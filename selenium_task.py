from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

from time import sleep
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active


# Set options for not prompting DevTools information
options = Options()
options.add_experimental_option("excludeSwitches", ["enable-logging"])

print("testing started")
driver = webdriver.Chrome(options=options)

driver.get("https://www.saucedemo.com/")
sleep(3)

# Find element using element's id attribute
driver.find_element(By.ID, "user-name").send_keys("standard_user")
driver.find_element(By.ID, "password").send_keys("secret_sauce")
driver.find_element(By.ID, "login-button").click()
sleep(5)

text = driver.find_element(By.CLASS_NAME, "title").text

# Check if login was successful 
assert "products" in text.lower()

print("TEST PASSED : LOGIN SUCCESSFUL")

# Get the Title, Desc and Price

print("testing get details")
inventory_item_names = driver.find_elements(By.CLASS_NAME, "inventory_item_name")
inventory_item_descs = driver.find_elements(By.CLASS_NAME, "inventory_item_desc")
inventory_item_prices = driver.find_elements(By.CLASS_NAME, "inventory_item_price")

# Write Workbook Column Headers
c1 = sheet.cell(row=1, column=1)
c1.value = "Title"

c2 = sheet.cell(row=1, column=2)
c2.value = "Description"

c3 = sheet.cell(row=1, column=3)
c3.value = "Price"



# Writes the data into the Workbook 
for i in range(len(inventory_item_names)):
    for j, item in enumerate([inventory_item_names, inventory_item_descs, inventory_item_prices], start=1):
        cell = sheet.cell(row=i+2, column=j)
        cell.value = item[i].text
        
         
print("TEST PASSED : GET DETAILS")


# Save the Workbook 
wb.save(r"C:\Users\User\Desktop\demo.xlsx") 

sleep(5)

# Close the driver
driver.quit()